#!/usr/bin/env python3
import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[2]
TAB_DIR = ROOT / "results" / "paper" / "final" / "tables"
OUT_XLSX = TAB_DIR / "supplementary_tables.xlsx"
ORDER = TAB_DIR / "table_order.tsv"

TABLES = [
    ("Table S1", TAB_DIR / "TableS1_main_binary_operating_characteristics.tsv"),
    ("Table S2", TAB_DIR / "TableS2_cohort_required_n.tsv"),
    ("Table S3", TAB_DIR / "TableS3_enrichment_summary.tsv"),
    ("Table S4", TAB_DIR / "TableS4_ordinal_binary_metrics_bias_key_n.tsv"),
    ("Table S5", TAB_DIR / "TableS5_ordinal_type1.tsv"),
    ("Table S6", TAB_DIR / "TableS6_ordinal_power_bias_tradeoff_key_n.tsv"),
    ("Table S7", TAB_DIR / "TableS7_ordinal_PO_nonPO_calibration.tsv"),
    ("Table S8", TAB_DIR / "TableS8_ordinal_PO_vs_nonPO_power_bias.tsv"),
]


def read_tsv(path: Path):
    with path.open("r", encoding="utf-8", newline="") as f:
        return list(csv.reader(f, delimiter="\t"))


def autosize(ws):
    for idx, column_cells in enumerate(ws.columns, start=1):
        values = [str(c.value) if c.value is not None else "" for c in column_cells]
        width = min(max(len(v) for v in values) + 2, 40)
        ws.column_dimensions[get_column_letter(idx)].width = width


def apply_arial(ws):
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                if cell.font and cell.font.bold:
                    cell.font = Font(name="Arial", bold=True)
                else:
                    cell.font = Font(name="Arial")


def main():
    wb = Workbook()
    ws = wb.active
    ws.title = "README"
    ws["A1"] = "Supplementary tables workbook"
    ws["A1"].font = Font(bold=True)
    ws.append([])
    ws.append(["table_id", "path", "purpose"])

    if ORDER.exists():
        with ORDER.open("r", encoding="utf-8", newline="") as f:
            r = csv.reader(f, delimiter="\t")
            next(r, None)
            for row in r:
                if row and row[0].startswith("Table S"):
                    ws.append(row)
    else:
        ws.append(["(missing)", str(ORDER), "table_order.tsv not found"])

    ws.append([])
    ws.append(["Note", "Blank proportional-bias cells indicate a true effect of zero in the null subgroup."])
    autosize(ws)
    apply_arial(ws)

    for name, path in TABLES:
        if not path.exists():
            continue
        sheet_name = name if len(name) <= 31 else name[:31]
        ws = wb.create_sheet(title=sheet_name)
        ws["A1"] = name
        ws["A1"].font = Font(bold=True)
        ws["A2"] = path.name
        rows = read_tsv(path)
        start_row = 4
        for r_idx, row in enumerate(rows, start=start_row):
            for c_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == start_row:
                    cell.font = Font(bold=True)
        ws.freeze_panes = f"A{start_row + 1}"
        autosize(ws)
        apply_arial(ws)

    wb.save(OUT_XLSX)
    print(f"Wrote {OUT_XLSX}")


if __name__ == "__main__":
    main()
